from openpyxl import load_workbook
import matplotlib.pyplot as plt


def read_excel_range() -> list:
    wb = load_workbook('tab1-zpl_08-2023.xlsx')
    sheet = wb.active

    values = []
    for row in range(8, 40):
        cell_value = sheet.cell(row=row, column=2).value
        values.append(cell_value)
    return values


def draw_graf(data):
    years = list(range(1991, 2023))
    salaries = data
    plt.figure(figsize=(12, 8))
    plt.bar(years, salaries, label="Зарплата")
    plt.xlabel('Год')
    plt.ylabel('Рублей (до 1998г. – тыс. рублей)')
    plt.title("""Среднемесячная номинальная начисленная
заработная плата работников по полному кругу организаций в целом по экономике
Российской Федерации в 1991-2023гг.
    """)
    plt.legend()
    plt.show()


def main():
    data = read_excel_range()
    draw_graf(data)


if __name__ == '__main__':
    main()
